"""
map_intake_to_import.py
=======================
Reads the filled device_intake_form.xlsx, applies all mapping logic
from mapping.yaml, and produces device_import_ready.xlsx — the same
format as the HTML form export, ready to feed into netbox_import.py.

Usage:
    python map_intake_to_import.py
    python map_intake_to_import.py --input my_inventory.xlsx
    python map_intake_to_import.py --input my_inventory.xlsx --output ready.xlsx

Requires:
    pip install openpyxl pyyaml pandas
"""

import os
import sys
import yaml
import argparse
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

MAPPING_FILE   = "mapping.yaml"
DEFAULT_INPUT  = "device_intake_form.xlsx"
DEFAULT_OUTPUT = "device_import_ready.xlsx"
INTAKE_SHEET   = "Device Intake"

# ── Colours ────────────────────────────────────────────────────────
BLUE_DARK  = "1F4E79"
BLUE_MED   = "2E75B6"
BLUE_LIGHT = "D6E4F0"
GREEN_DARK = "375623"
GREEN_LT   = "EBF1DE"
ORANGE     = "C55A11"
ORANGE_LT  = "FFF2CC"
GRAY       = "F2F2F2"
WHITE      = "FFFFFF"
RED_LT     = "FDEEEE"


def load_mapping() -> dict:
    if not os.path.exists(MAPPING_FILE):
        print(f"[ERROR] {MAPPING_FILE} not found in current directory.")
        sys.exit(1)
    with open(MAPPING_FILE, encoding="utf-8") as f:
        return yaml.safe_load(f)


def clean(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in ("nan", "none", "n/a", "-") else s


# ═══════════════════════════════════════════════════════════════════
# Mapping engine — mirrors form HTML logic, driven by mapping.yaml
# ═══════════════════════════════════════════════════════════════════

def apply_mapping(row: dict, mapping: dict) -> dict:
    asset_type       = clean(row.get("asset_type", ""))
    os_val           = clean(row.get("os", ""))
    server_role      = clean(row.get("server_role", ""))
    workstation_type = clean(row.get("workstation_type", ""))
    appliance_type   = clean(row.get("appliance_type", ""))
    internet_facing  = clean(row.get("internet_facing", ""))
    sensitive_data   = clean(row.get("sensitive_data", ""))

    # Resolve specific_product — dropdown takes priority, freetext as fallback
    product_dropdown = clean(row.get("specific_product", ""))
    product_freetext = clean(row.get("product_freetext", ""))
    specific_product = product_freetext if (not product_dropdown or product_dropdown.lower() == "other") else product_dropdown

    # Start with defaults
    defaults = mapping.get("defaults", {})
    result = {
        "hostname":         clean(row.get("hostname", "")),
        "ip_address":       clean(row.get("ip_address", "")),
        "mac_address":      clean(row.get("mac_address", "")),
        "asset_type":       asset_type,
        "os":               os_val,
        "environment":      clean(row.get("environment", "")),
        "sensitive_data":   sensitive_data,
        "site":             clean(row.get("site", "")),
        "department":       clean(row.get("department", "")),
        "server_role":      server_role,
        "workstation_type": workstation_type,
        "specific_product": specific_product,
        "internet_facing":  internet_facing,
        "appliance_type":   appliance_type,
        "appliance_vendor": clean(row.get("appliance_vendor", "")),
        "syslog_configured":clean(row.get("syslog_configured", "")),
        "notes":            clean(row.get("notes", "")),
        # Mapped fields — populated below
        "device_function":          "",
        "cis_os_benchmark":         "",
        "cis_os_profile":           defaults.get("cis_os_profile", "L1"),
        "cis_app_benchmark":        "",
        "sigma_product":            "",
        "sigma_log_categories":     "",
        "business_app_log_path":    defaults.get("business_app_log_path", ""),
        "cis_os_status":            defaults.get("cis_os_status", "NOT-STARTED"),
        "cis_app_status":           defaults.get("cis_app_status", "NOT-STARTED"),
        "log_collection_status":    defaults.get("log_collection_status", "NOT-CONFIGURED"),
        "log_collector":            defaults.get("log_collector", "NONE"),
        "hardening_exception_notes":"",
        "import_status":            "Ready",
    }

    # ── OS benchmark ──────────────────────────────────────────────
    os_map = mapping.get("os_to_cis_benchmark", {})
    result["cis_os_benchmark"] = os_map.get(os_val, "OTHER" if os_val else "")

    # ── Asset-type routing ────────────────────────────────────────
    if asset_type == "Server":
        server_roles = mapping.get("server_roles", {})
        role = server_roles.get(server_role, server_roles.get("OTHER", {}))

        result["device_function"]      = role.get("device_function", "OTHER")
        result["cis_app_benchmark"]    = role.get("cis_app_benchmark", "OTHER")
        result["sigma_product"]        = list(role.get("sigma_product", []))
        result["sigma_log_categories"] = list(role.get("sigma_log_categories", []))

        if role.get("flag_log_path"):
            result["business_app_log_path"] = "PENDING — Blue Team to define"

        if role.get("import_status"):
            result["import_status"] = role["import_status"]

        # Product overrides (Q2)
        if role.get("product_overrides") and specific_product:
            k = specific_product.lower()
            p_cis = mapping.get("product_to_cis_app_benchmark", {})
            p_sig = mapping.get("product_to_sigma", {})
            if k in p_cis:
                result["cis_app_benchmark"] = p_cis[k]
            if k in p_sig:
                result["sigma_product"] = list(p_sig[k])

    elif asset_type == "Workstation":
        wrk_default = mapping.get("workstation_default", "WRK-STD")
        wtype = workstation_type if workstation_type else wrk_default
        wrk_roles = mapping.get("workstation_roles", {})
        role = wrk_roles.get(wtype, wrk_roles.get(wrk_default, {}))

        result["device_function"]      = role.get("device_function", "WRK-STD")
        result["cis_app_benchmark"]    = role.get("cis_app_benchmark", "NA")
        result["sigma_product"]        = list(role.get("sigma_product", ["windows"]))
        result["sigma_log_categories"] = list(role.get("sigma_log_categories", []))

    elif asset_type == "Appliance":
        apl_roles = mapping.get("appliance_roles", {})
        role = apl_roles.get(appliance_type, apl_roles.get("OTHER", {}))

        result["device_function"]      = role.get("device_function", "OTHER")
        result["cis_os_benchmark"]     = role.get("cis_os_benchmark", "VENDOR-OS")
        result["cis_app_benchmark"]    = role.get("cis_app_benchmark", "VENDOR-GUIDE")
        result["sigma_product"]        = list(role.get("sigma_product", ["custom"]))
        result["sigma_log_categories"] = list(role.get("sigma_log_categories", []))
        result["log_collector"]        = role.get("log_collector", "SYSLOG")

        if role.get("import_status"):
            result["import_status"] = role["import_status"]

    else:
        # Unknown asset type
        result["device_function"] = "OTHER"
        result["import_status"]   = "REVIEW NEEDED"

    # ── Escalation rules ─────────────────────────────────────────
    esc = mapping.get("escalation", {})
    esc_fields = esc.get("fields", [])
    esc_values = esc.get("values", [])
    triggered  = any(
        clean(row.get(f, "")) in esc_values
        for f in esc_fields
    )
    if triggered:
        for k, v in esc.get("overrides", {}).items():
            result[k] = v
        for cat in esc.get("append_sigma_categories", []):
            if cat not in result["sigma_log_categories"]:
                result["sigma_log_categories"].append(cat)

    # ── Stringify multiselect lists ───────────────────────────────
    result["sigma_product"]        = ",".join(result["sigma_product"])
    result["sigma_log_categories"] = ",".join(result["sigma_log_categories"])

    return result


# ═══════════════════════════════════════════════════════════════════
# Excel output
# ═══════════════════════════════════════════════════════════════════

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def fnt(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def aln(h="left"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)


# Column definitions for the import Excel
# (field, label, width, group)  group: form / mapped / default / meta
IMPORT_COLUMNS = [
    # Form fields
    ("hostname",                "Hostname",              22, "form"),
    ("ip_address",              "IP Address",            18, "form"),
    ("mac_address",             "MAC Address",           20, "form"),
    ("asset_type",              "Asset Type",            14, "form"),
    ("os",                      "Operating System",      20, "form"),
    ("environment",             "Environment",           14, "form"),
    ("sensitive_data",          "Sensitive Data",        14, "form"),
    ("site",                    "Site",                  16, "form"),
    ("department",              "Department",            18, "form"),
    ("server_role",             "Server Role",           14, "form"),
    ("workstation_type",        "Workstation Type",      16, "form"),
    ("specific_product",        "Specific Product",      20, "form"),
    ("internet_facing",         "Internet Facing",       14, "form"),
    ("appliance_type",          "Appliance Type",        14, "form"),
    ("appliance_vendor",        "Vendor & Model",        22, "form"),
    ("syslog_configured",       "Syslog Forwarding",     14, "form"),
    # Auto-mapped NetBox custom fields
    ("device_function",         "device_function",       18, "mapped"),
    ("cis_os_benchmark",        "cis_os_benchmark",      18, "mapped"),
    ("cis_os_profile",          "cis_os_profile",        14, "mapped"),
    ("cis_app_benchmark",       "cis_app_benchmark",     18, "mapped"),
    ("sigma_product",           "sigma_product",         26, "mapped"),
    ("sigma_log_categories",    "sigma_log_categories",  34, "mapped"),
    ("business_app_log_path",   "business_app_log_path", 24, "mapped"),
    # Defaults
    ("cis_os_status",           "cis_os_status",         16, "default"),
    ("cis_app_status",          "cis_app_status",        16, "default"),
    ("log_collection_status",   "log_collection_status", 20, "default"),
    ("log_collector",           "log_collector",         16, "default"),
    ("hardening_exception_notes","hardening_exception_notes",24,"default"),
    # Meta
    ("notes",                   "Notes",                 28, "meta"),
    ("import_status",           "import_status",         16, "meta"),
]

GROUP_COLORS = {
    "form":    (BLUE_MED,   BLUE_LIGHT),
    "mapped":  (GREEN_DARK, GREEN_LT),
    "default": ("595959",   GRAY),
    "meta":    (ORANGE,     ORANGE_LT),
}


def write_import_excel(records: list, output_path: str, mapping: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Device Onboarding"

    ncols = len(IMPORT_COLUMNS)

    # Row 1 — title
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"] = f"SOC Device Import — Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].fill = fill(BLUE_DARK)
    ws["A1"].font = fnt(bold=True, color=WHITE, size=12)
    ws["A1"].alignment = aln("center")
    ws.row_dimensions[1].height = 22

    # Row 2 — group headers
    ws.row_dimensions[2].height = 14
    prev_group = None
    group_start = 1
    for i, (field, label, width, group) in enumerate(IMPORT_COLUMNS, start=1):
        if group != prev_group:
            group_start = i
            prev_group = group
        hdr_color, _ = GROUP_COLORS[group]
        cell = ws.cell(row=2, column=i)
        cell.fill = fill(hdr_color)
        cell.font = fnt(bold=True, color=WHITE, size=8)
        cell.alignment = aln("center")
        cell.border = bdr()
        if i == group_start:
            cell.value = group.upper()

    # Row 3 — column headers
    ws.row_dimensions[3].height = 30
    for i, (field, label, width, group) in enumerate(IMPORT_COLUMNS, start=1):
        _, lt_color = GROUP_COLORS[group]
        cell = ws.cell(row=3, column=i, value=label)
        cell.fill      = fill(lt_color)
        cell.font      = fnt(bold=True, color=BLUE_DARK if group == "form" else GREEN_DARK if group == "mapped" else "595959", size=9)
        cell.alignment = aln("center")
        cell.border    = bdr()
        ws.column_dimensions[get_column_letter(i)].width = width

    # Data rows
    review_count = 0
    for row_idx, rec in enumerate(records, start=4):
        is_review = rec.get("import_status", "") == "REVIEW NEEDED"
        row_bg    = RED_LT if is_review else (WHITE if row_idx % 2 == 0 else GRAY)
        if is_review:
            review_count += 1

        ws.row_dimensions[row_idx].height = 15
        for col_idx, (field, label, width, group) in enumerate(IMPORT_COLUMNS, start=1):
            val  = rec.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill(row_bg)
            cell.font      = fnt(size=10, bold=(field == "hostname"))
            cell.alignment = aln()
            cell.border    = bdr()

    ws.freeze_panes = "A4"

    # README sheet
    rs = wb.create_sheet("README")
    readme_rows = [
        ("SOC Device Import — README", True, BLUE_DARK, WHITE),
        (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Source: {DEFAULT_INPUT}", False, BLUE_MED, WHITE),
        ("", False, WHITE, "000000"),
        ("NEXT STEP", True, BLUE_MED, WHITE),
        ("Run: python netbox_import.py --file device_import_ready.xlsx --dry-run", False, WHITE, "000000"),
        ("Then: python netbox_import.py --file device_import_ready.xlsx", False, WHITE, "000000"),
        ("", False, WHITE, "000000"),
        ("COLUMN GROUPS", True, BLUE_MED, WHITE),
        ("FORM (blue)    — original data from intake sheet", False, BLUE_LIGHT, BLUE_DARK),
        ("MAPPED (green) — auto-derived NetBox custom fields from mapping.yaml", False, GREEN_LT, GREEN_DARK),
        ("DEFAULT (gray) — starting status values, updated later by tools", False, GRAY, "595959"),
        ("META (orange)  — import_status: Ready = import | REVIEW NEEDED = fix first", False, ORANGE_LT, ORANGE),
        ("", False, WHITE, "000000"),
        ("REVIEW NEEDED rows", True, BLUE_MED, WHITE),
        ("Rows highlighted in red have import_status = REVIEW NEEDED.", False, RED_LT, "7B2C2C"),
        ("These have device_function = OTHER — set the correct function manually, change import_status to Ready, re-run.", False, RED_LT, "7B2C2C"),
    ]
    for i, (text, bold, bg, fg) in enumerate(readme_rows, start=1):
        rs.merge_cells(f"A{i}:C{i}")
        cell = rs.cell(row=i, column=1, value=text)
        cell.fill = fill(bg)
        cell.font = fnt(bold=bold, color=fg, size=10)
        cell.alignment = aln()
        cell.border = bdr()
        rs.row_dimensions[i].height = 18
    rs.column_dimensions["A"].width = 80

    wb.save(output_path)
    return review_count


# ═══════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="SOC Intake to Import Excel Mapper")
    parser.add_argument("--input",  default=DEFAULT_INPUT,  help=f"Intake Excel file (default: {DEFAULT_INPUT})")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help=f"Output Excel file (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--sheet",  default=INTAKE_SHEET,   help=f"Sheet name to read (default: '{INTAKE_SHEET}')")
    args = parser.parse_args()

    print("=" * 60)
    print("  SOC Intake → Import Excel Mapper")
    print(f"  Input:   {args.input}")
    print(f"  Output:  {args.output}")
    print("=" * 60)

    mapping = load_mapping()
    print(f"  Mapping: {MAPPING_FILE} loaded")

    # Read intake Excel
    if not os.path.exists(args.input):
        print(f"\n[ERROR] Input file not found: {args.input}")
        print(f"        Run: python generate_intake_excel.py")
        print(f"        Fill the intake form, then re-run this script.")
        sys.exit(1)

    try:
        df = pd.read_excel(args.input, sheet_name=args.sheet, dtype=str, header=2)
        # Drop completely empty rows
        df = df.dropna(how="all")
        # Strip leading * from header names (required field markers)
        df.columns = [c.lstrip("* ").strip() for c in df.columns]
        # Normalize column names to field names
        col_rename = {
            "Hostname":               "hostname",
            "IP Address":             "ip_address",
            "MAC Address":            "mac_address",
            "Asset Type":             "asset_type",
            "Operating System":       "os",
            "Environment":            "environment",
            "Sensitive Data?":        "sensitive_data",
            "Site":                   "site",
            "Department / Owner":     "department",
            "Server Role (Q1)":       "server_role",
            "Product / App (Q2)":     "specific_product",
            "If not listed — App":    "product_freetext",
            "Internet Facing? (Q3)":  "internet_facing",
            "Workstation Type":       "workstation_type",
            "Appliance Type":         "appliance_type",
            "Vendor & Model":         "appliance_vendor",
            "Syslog Forwarding?":     "syslog_configured",
            "Notes / Exceptions":     "notes",
        }
        df = df.rename(columns=col_rename)
        df = df.fillna("")
        print(f"  Intake:  {len(df)} rows read from '{args.sheet}'\n")
    except Exception as e:
        print(f"\n[ERROR] Cannot read intake file: {e}")
        sys.exit(1)

    # Process each row
    records    = []
    skipped    = 0
    for _, row in df.iterrows():
        if not clean(row.get("hostname", "")):
            skipped += 1
            continue
        mapped = apply_mapping(row.to_dict(), mapping)
        records.append(mapped)

    print(f"  Processed: {len(records)} devices  |  {skipped} skipped (no hostname)\n")

    # Status summary
    status_counts = {}
    for r in records:
        s = r.get("import_status", "Ready")
        status_counts[s] = status_counts.get(s, 0) + 1

    for status, count in status_counts.items():
        icon = "✓" if status == "Ready" else "⚠"
        print(f"  {icon}  {status}: {count}")

    # Write output
    review_count = write_import_excel(records, args.output, mapping)
    print(f"\n  ✓ Written: {args.output}")

    if review_count:
        print(f"\n  ⚠  {review_count} device(s) marked REVIEW NEEDED (highlighted red).")
        print(f"     Open {args.output}, set device_function manually for those rows,")
        print(f"     change import_status to 'Ready', then run netbox_import.py.")

    print(f"\n  Next step:")
    print(f"    python netbox_import.py --file {args.output} --dry-run")
    print("=" * 60)


if __name__ == "__main__":
    main()
