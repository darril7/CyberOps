"""
generate_intake_excel.py
========================
Generates device_intake_form.xlsx — an Excel version of the onboarding form.

Use this when you already have a device inventory in Excel and want to
copy-paste it into a structured sheet with dropdown validation.
Fill the intake form, then run map_intake_to_import.py to produce the
NetBox-ready import Excel.

Usage:
    python generate_intake_excel.py

Requires:
    pip install openpyxl pyyaml
"""

import os
import sys
import yaml
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MAPPING_FILE = "mapping.yaml"
OUTPUT_FILE  = "device_intake_form.xlsx"

# ── Colours ────────────────────────────────────────────────────────
BLUE_DARK   = "1F4E79"
BLUE_MED    = "2E75B6"
BLUE_LIGHT  = "D6E4F0"
GREEN_DARK  = "375623"
GREEN_LIGHT = "EBF1DE"
ORANGE      = "C55A11"
ORANGE_LT   = "FFF2CC"
YELLOW      = "FFFF00"
GRAY        = "F2F2F2"
WHITE       = "FFFFFF"
RED_LIGHT   = "FDEEEE"


def load_mapping() -> dict:
    if not os.path.exists(MAPPING_FILE):
        print(f"[ERROR] {MAPPING_FILE} not found.")
        sys.exit(1)
    with open(MAPPING_FILE, encoding="utf-8") as f:
        return yaml.safe_load(f)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(cell, bg=BLUE_DARK, fg=WHITE, bold=True):
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=10)
    cell.alignment = center()
    cell.border = border()


def style_subheader(cell, bg=BLUE_LIGHT):
    cell.fill = fill(bg)
    cell.font = font(bold=True, color=BLUE_DARK, size=9)
    cell.alignment = center()
    cell.border = border()


def style_required(cell):
    cell.fill = fill(ORANGE_LT)
    cell.font = font(bold=True, color=ORANGE, size=9)
    cell.alignment = center()
    cell.border = border()


def style_optional(cell):
    cell.fill = fill(GRAY)
    cell.font = font(bold=False, color="595959", size=9)
    cell.alignment = center()
    cell.border = border()


def style_data(cell, bg=WHITE):
    cell.fill = fill(bg)
    cell.font = font(size=10)
    cell.alignment = left()
    cell.border = border()


def add_dropdown(ws, col_letter, start_row, end_row, formula, title, prompt):
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showInputMessage=True,
        promptTitle=title,
        prompt=prompt,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error="Please select a value from the dropdown list.",
    )
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def build_choices_sheet(wb, mapping):
    """
    Hidden sheet that stores all dropdown list values.
    Named ranges reference columns here.
    """
    cs = wb.create_sheet("_choices")
    cs.sheet_state = "hidden"

    col = 1

    def write_list(header, values):
        nonlocal col
        ltr = get_column_letter(col)
        cs[f"{ltr}1"] = header
        cs[f"{ltr}1"].font = Font(name="Arial", bold=True, size=9)
        for i, v in enumerate(values, start=2):
            cs[f"{ltr}{i}"] = v
            cs[f"{ltr}{i}"].font = Font(name="Arial", size=9)
        end_row = 1 + len(values)
        ref = f"_choices!${ltr}$2:${ltr}${end_row}"
        col += 1
        return ref

    refs = {}

    # Asset type
    refs["asset_type"] = write_list("asset_type", ["Server", "Workstation", "Appliance"])

    # OS
    os_list = list(mapping.get("os_to_cis_benchmark", {}).keys())
    refs["os"] = write_list("os", os_list)

    # Environment
    refs["environment"] = write_list("environment", ["Production", "Development", "Test", "DMZ"])

    # Yes/No
    refs["yesno"] = write_list("yes_no", ["Yes", "No"])

    # Server roles
    srv_codes = list(mapping.get("server_roles", {}).keys())
    refs["server_role"] = write_list("server_role", srv_codes)

    # Products
    products = list({**mapping.get("product_to_cis_app_benchmark", {}),
                     **mapping.get("product_to_sigma", {})}.keys())
    products = sorted(set(p.title() for p in products)) + ["Other"]
    refs["product"] = write_list("product", products)

    # Workstation types
    wrk_codes = list(mapping.get("workstation_roles", {}).keys())
    refs["workstation_type"] = write_list("workstation_type", wrk_codes)

    # Appliance types
    apl_codes = list(mapping.get("appliance_roles", {}).keys())
    refs["appliance_type"] = write_list("appliance_type", apl_codes)

    # Syslog
    refs["syslog"] = write_list("syslog", ["Yes", "No", "Pending"])

    # Set column widths
    for i in range(1, col):
        cs.column_dimensions[get_column_letter(i)].width = 22

    return refs


def build_intake_sheet(wb, mapping, refs):
    ws = wb.active
    ws.title = "Device Intake"

    # ── Column definitions ─────────────────────────────────────────
    # (field_name, display_label, width, required, type, choices_ref_key or None)
    COLUMNS = [
        # Basic info
        ("hostname",          "Hostname",             22, True,  "text",     None),
        ("ip_address",        "IP Address",           18, False, "text",     None),
        ("mac_address",       "MAC Address",          20, False, "text",     None),
        ("asset_type",        "Asset Type",           16, True,  "dropdown", "asset_type"),
        ("os",                "Operating System",     22, False, "dropdown", "os"),
        ("environment",       "Environment",          15, False, "dropdown", "environment"),
        ("sensitive_data",    "Sensitive Data?",      16, False, "dropdown", "yesno"),
        ("department",        "Department / Owner",   22, False, "text",     None),
        # Server fields
        ("server_role",       "Server Role (Q1)",     22, False, "dropdown", "server_role"),
        ("specific_product",  "Product / App (Q2)",   22, False, "dropdown", "product"),
        ("product_freetext",  "If not listed — App",  24, False, "text",     None),
        ("internet_facing",   "Internet Facing? (Q3)",18, False, "dropdown", "yesno"),
        # Workstation
        ("workstation_type",  "Workstation Type",     20, False, "dropdown", "workstation_type"),
        # Appliance
        ("appliance_type",    "Appliance Type",       18, False, "dropdown", "appliance_type"),
        ("appliance_vendor",  "Vendor & Model",       24, False, "text",     None),
        ("syslog_configured", "Syslog Forwarding?",   18, False, "dropdown", "syslog"),
        # Notes
        ("notes",             "Notes / Exceptions",   30, False, "text",     None),
    ]

    DATA_START_ROW = 4
    DATA_END_ROW   = 503   # 500 rows of data

    # ── Row 1: title banner ────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    ws["A1"] = "SOC Program — Device Intake Form"
    ws["A1"].fill  = fill(BLUE_DARK)
    ws["A1"].font  = Font(name="Arial", bold=True, color=WHITE, size=13)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 26

    # ── Row 2: subtitle ───────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
    ws["A2"] = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d')}  |  "
        f"Copy-paste your inventory here, fill required fields, then run: "
        f"python map_intake_to_import.py"
    )
    ws["A2"].fill      = fill(BLUE_MED)
    ws["A2"].font      = Font(name="Arial", color=WHITE, size=9, italic=True)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 16

    # ── Row 3: column headers ─────────────────────────────────────
    ws.row_dimensions[3].height = 36
    for i, (field, label, width, required, ftype, _) in enumerate(COLUMNS, start=1):
        col_ltr = get_column_letter(i)
        cell = ws[f"{col_ltr}3"]
        cell.value = f"{'* ' if required else ''}{label}"
        if required:
            style_required(cell)
        else:
            style_subheader(cell)
        ws.column_dimensions[col_ltr].width = width

    # ── Data rows: styling + dropdowns ───────────────────────────
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        bg = WHITE if row % 2 == 0 else GRAY
        for i in range(1, len(COLUMNS) + 1):
            cell = ws[f"{get_column_letter(i)}{row}"]
            style_data(cell, bg)
        ws.row_dimensions[row].height = 16

    # ── Add dropdowns ─────────────────────────────────────────────
    for i, (field, label, width, required, ftype, ref_key) in enumerate(COLUMNS, start=1):
        if ftype == "dropdown" and ref_key and ref_key in refs:
            col_ltr = get_column_letter(i)
            formula = f'"{refs[ref_key]}"'  # openpyxl accepts range string
            # Use the actual range reference
            formula = refs[ref_key]
            add_dropdown(
                ws, col_ltr, DATA_START_ROW, DATA_END_ROW,
                formula=formula,
                title=label,
                prompt=f"Select a value for {label}"
            )

    # ── Freeze panes ──────────────────────────────────────────────
    ws.freeze_panes = f"A{DATA_START_ROW}"

    # ── Section color bands in header row ─────────────────────────
    # Visual grouping — basic / server / workstation / appliance / notes
    sections = [
        (1,  7,  BLUE_MED,   "BASIC INFO"),
        (8,  11, GREEN_DARK,  "SERVER (fill Q1-Q3 for servers)"),
        (12, 12, "5B5EA6",    "WORKSTATION"),
        (13, 15, ORANGE,      "APPLIANCE"),
        (16, 16, "595959",    "NOTES"),
    ]
    # Add a row 2.5 — section labels above headers
    # Insert a section label row by using row 3 merged sub-labels
    # (Already done via header colour, keep it clean)

    return ws, COLUMNS, DATA_START_ROW, DATA_END_ROW


def build_legend_sheet(wb, mapping):
    ls = wb.create_sheet("Legend & Instructions")

    def write(row, col, text, bold=False, bg=None, color="000000", size=10, merge_to=None):
        cell = ls.cell(row=row, column=col, value=text)
        cell.font      = Font(name="Arial", bold=bold, color=color, size=size)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border    = border()
        if bg:
            cell.fill = fill(bg)
        if merge_to:
            ls.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
        return cell

    r = 1
    write(r, 1, "SOC Device Intake Form — Legend & Instructions",
          bold=True, bg=BLUE_DARK, color=WHITE, size=13, merge_to=4)
    ls.row_dimensions[r].height = 24
    r += 2

    # Instructions
    write(r, 1, "HOW TO USE", bold=True, bg=BLUE_MED, color=WHITE, merge_to=4)
    r += 1
    steps = [
        "1.  Copy-paste your existing inventory into the 'Device Intake' sheet starting at row 4.",
        "2.  Fill at minimum: Hostname, Asset Type, and OS for every device.",
        "3.  For Servers: fill Server Role (Q1). If the role has a specific product (DB, Web, Mail), fill Product (Q2). Fill Internet Facing (Q3).",
        "4.  For Workstations: fill Workstation Type (or leave blank for Standard User default).",
        "5.  For Appliances: fill Appliance Type and Vendor & Model.",
        "6.  Use the dropdowns — they validate against your mapping.yaml choices.",
        "7.  Save the file, then run:  python map_intake_to_import.py  to generate the NetBox import Excel.",
    ]
    for step in steps:
        write(r, 1, step, merge_to=4)
        ls.row_dimensions[r].height = 20
        r += 1
    r += 1

    # Column guide
    write(r, 1, "COLUMN GUIDE", bold=True, bg=BLUE_MED, color=WHITE, merge_to=4)
    r += 1
    write(r, 1, "Column", bold=True, bg=BLUE_LIGHT)
    write(r, 2, "Required?", bold=True, bg=BLUE_LIGHT)
    write(r, 3, "Type", bold=True, bg=BLUE_LIGHT)
    write(r, 4, "Description", bold=True, bg=BLUE_LIGHT)
    r += 1

    col_guide = [
        ("Hostname",              "YES",  "Text",     "Device hostname — primary key for NetBox. Must be unique."),
        ("IP Address",            "no",   "Text",     "Management IP address."),
        ("MAC Address",           "no",   "Text",     "Primary/management interface MAC address. Format: AA:BB:CC:DD:EE:FF. Used to create a Management interface in NetBox."),
        ("Asset Type",            "YES",  "Dropdown", "Server / Workstation / Appliance — determines which other fields apply."),
        ("Operating System",      "no",   "Dropdown", "OS version — drives the CIS OS benchmark auto-mapping."),
        ("Environment",           "no",   "Dropdown", "Production / Development / Test / DMZ."),
        ("Sensitive Data?",       "no",   "Dropdown", "Yes escalates CIS hardening profile to Level 2."),
        ("Department / Owner",    "no",   "Text",     "Team or person responsible. Mainly for workstations."),
        ("Server Role (Q1)",      "no",   "Dropdown", "Required for servers. Drives CIS app benchmark + Sigma profile."),
        ("Product / App (Q2)",    "no",   "Dropdown", "Required when role has a specific product (DB, Web, Mail)."),
        ("If not listed — App",   "no",   "Text",     "Free text for business apps not in the product dropdown."),
        ("Internet Facing? (Q3)", "no",   "Dropdown", "Yes escalates CIS to L2 and adds network_connection Sigma category."),
        ("Workstation Type",      "no",   "Dropdown", "WRK-STD / WRK-DEV / WRK-PAW. Defaults to WRK-STD if blank."),
        ("Appliance Type",        "no",   "Dropdown", "APL-FW / APL-SW / APL-WAF / APL-IDS / APL-PROXY."),
        ("Vendor & Model",        "no",   "Text",     "Appliance vendor and model e.g. Fortinet FortiGate 100F."),
        ("Syslog Forwarding?",    "no",   "Dropdown", "For appliances — whether syslog is configured."),
        ("Notes / Exceptions",    "no",   "Text",     "Anything the security team should know about this device."),
    ]

    for cname, req, ctype, desc in col_guide:
        bg = ORANGE_LT if req == "YES" else WHITE
        write(r, 1, cname,  bg=bg)
        write(r, 2, req,    bg=bg, bold=(req=="YES"), color=ORANGE if req=="YES" else "595959")
        write(r, 3, ctype,  bg=bg)
        write(r, 4, desc,   bg=bg)
        r += 1
    r += 1

    # Mapping guide
    write(r, 1, "MAPPING REFERENCE — What gets auto-derived", bold=True, bg=GREEN_DARK, color=WHITE, merge_to=4)
    r += 1
    write(r, 1, "Form Input", bold=True, bg=GREEN_LIGHT)
    write(r, 2, "→  NetBox Field", bold=True, bg=GREEN_LIGHT)
    write(r, 3, "→  CIS Benchmark", bold=True, bg=GREEN_LIGHT)
    write(r, 4, "→  Sigma Product", bold=True, bg=GREEN_LIGHT)
    r += 1

    mapping_ref = [
        ("OS = Windows Server 2022",        "cis_os_benchmark",                      "CIS-WS2022",            "windows"),
        ("OS = Ubuntu 22.04",               "cis_os_benchmark",                      "CIS-UB2204",            "linux"),
        ("Role = SRV-DC",                   "device_function + cis_app_benchmark",   "CIS Active Directory",  "windows, active_directory"),
        ("Role = SRV-DB + Product = MySQL", "device_function + cis_app_benchmark",   "CIS MySQL",             "linux, mysql"),
        ("Role = SRV-WEB + Product = IIS",  "device_function + cis_app_benchmark",   "CIS IIS",               "windows, iis"),
        ("Role = SRV-APP (custom app)",     "device_function, flag_log_path=PENDING","Custom Baseline",       "custom"),
        ("Internet Facing = Yes",           "cis_os_profile = L2",                   "Escalate to L2",        "+ network_connection"),
        ("Appliance = APL-FW (Fortinet)",   "device_function, log_collector=SYSLOG", "Vendor Guide",          "fortinet, firewall"),
    ]
    for a, b, c, d in mapping_ref:
        write(r, 1, a)
        write(r, 2, b, color=BLUE_DARK)
        write(r, 3, c, color=GREEN_DARK)
        write(r, 4, d, color=ORANGE)
        r += 1

    ls.column_dimensions["A"].width = 35
    ls.column_dimensions["B"].width = 28
    ls.column_dimensions["C"].width = 25
    ls.column_dimensions["D"].width = 55
    ls.freeze_panes = "A3"


def main():
    print("=" * 55)
    print("  SOC Device Intake Form Generator")
    print(f"  Reading: {MAPPING_FILE}")
    print("=" * 55)

    mapping = load_mapping()
    print(f"  Mapping loaded — {len(mapping.get('server_roles',{}))} server roles, "
          f"{len(mapping.get('appliance_roles',{}))} appliance types\n")

    wb = Workbook()

    refs = build_choices_sheet(wb, mapping)
    build_intake_sheet(wb, mapping, refs)
    build_legend_sheet(wb, mapping)

    # Sheet order: intake first
    wb.move_sheet("Device Intake", offset=-wb.sheetnames.index("Device Intake"))

    wb.save(OUTPUT_FILE)
    print(f"  ✓ Written: {OUTPUT_FILE}")
    print(f"\n  Fill this sheet with your inventory, then run:")
    print(f"    python map_intake_to_import.py")
    print("=" * 55)


if __name__ == "__main__":
    main()
